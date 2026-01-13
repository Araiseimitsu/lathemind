"""
XLSX工程ファイルパーサー

シチズンマシナリー CINCOM L20 用工程管理表のパース処理
"""

from openpyxl import load_workbook
from typing import List, Dict, Optional
import logging

logger = logging.getLogger(__name__)

# 定数
SHEET_NAME = "加工工程管理表"
START_ROW = 12  # 12行目から開始（1-based）
ROW_SKIP = 2    # 1行おきなので2行ごと

# 正面側工程の列
FRONT_COLUMNS = {
    "correction": "A",  # 補正番号
    "tool": "C",        # ツール番号
    "name": "E",        # 工程名（工具名）
}

# 背面側工程の列
BACK_COLUMNS = {
    "correction": "S",  # 補正番号
    "tool": "U",        # ツール番号
    "name": "W",        # 工程名（工程番号）
}


def _parse_operations(
    sheet,
    columns: Dict[str, str],
    start_row: int = START_ROW,
    row_skip: int = ROW_SKIP,
    max_rows: int = 100
) -> List[Dict[str, str]]:
    """
    シートから工程データをパースする

    Args:
        sheet: openpyxlワークシートオブジェクト
        columns: 列名とセル列のマッピング {"correction": "A", "tool": "C", "name": "E"}
        start_row: 開始行（1-based）
        row_skip: 読み飛ばし行数
        max_rows: 最大読み取り行数

    Returns:
        工程データのリスト [{"correction": "A12", "tool": "T1", "name": "ZAGURI"}, ...]
    """
    operations = []

    for row_offset in range(0, max_rows):
        row_num = start_row + (row_offset * row_skip)

        # 各列の値を取得
        correction_cell = f"{columns['correction']}{row_num}"
        tool_cell = f"{columns['tool']}{row_num}"
        name_cell = f"{columns['name']}{row_num}"

        correction = sheet[correction_cell].value
        tool = sheet[tool_cell].value
        name = sheet[name_cell].value

        # すべて空の場合、終了とみなす
        if not correction and not tool and not name:
            break

        # 空の工程はスキップ
        if not name:
            continue

        operations.append({
            "correction": str(correction) if correction else "",
            "tool": str(tool) if tool else "",
            "name": str(name).strip() if name else ""
        })

    return operations


def parse_process_file(file_path: str) -> Dict[str, List[Dict[str, str]]]:
    """
    XLSX工程管理ファイルをパースする

    Args:
        file_path: XLSXファイルのパス

    Returns:
        パース結果 {"front_operations": [...], "back_operations": [...]}
    """
    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)

        # シート名を検索（完全一致または部分一致）
        sheet = None
        for sheet_name in wb.sheetnames:
            if SHEET_NAME in sheet_name:
                sheet = wb[sheet_name]
                break

        if sheet is None:
            # シートが見つからない場合は最初のシートを使用
            logger.warning(f"シート「{SHEET_NAME}」が見つかりません。最初のシートを使用します。")
            sheet = wb.active

        # 正面側工程をパース
        front_operations = _parse_operations(sheet, FRONT_COLUMNS)

        # 背面側工程をパース
        back_operations = _parse_operations(sheet, BACK_COLUMNS)

        wb.close()

        logger.info(
            f"パース完了: 正面側 {len(front_operations)}工程, "
            f"背面側 {len(back_operations)}工程"
        )

        return {
            "front_operations": front_operations,
            "back_operations": back_operations
        }

    except Exception as e:
        logger.error(f"XLSXパースエラー: {e}")
        raise


def parse_process_file_bytes(file_content: bytes) -> Dict[str, List[Dict[str, str]]]:
    """
    XLSX工程管理ファイルをパースする（bytes入力版）

    Args:
        file_content: XLSXファイルのバイナリデータ

    Returns:
        パース結果 {"front_operations": [...], "back_operations": [...]}
    """
    import io

    try:
        wb = load_workbook(filename=io.BytesIO(file_content), read_only=True, data_only=True)

        # シート名を検索（完全一致または部分一致）
        sheet = None
        for sheet_name in wb.sheetnames:
            if SHEET_NAME in sheet_name:
                sheet = wb[sheet_name]
                break

        if sheet is None:
            # シートが見つからない場合は最初のシートを使用
            logger.warning(f"シート「{SHEET_NAME}」が見つかりません。最初のシートを使用します。")
            sheet = wb.active

        # 正面側工程をパース
        front_operations = _parse_operations(sheet, FRONT_COLUMNS)

        # 背面側工程をパース
        back_operations = _parse_operations(sheet, BACK_COLUMNS)

        wb.close()

        logger.info(
            f"パース完了: 正面側 {len(front_operations)}工程, "
            f"背面側 {len(back_operations)}工程"
        )

        return {
            "front_operations": front_operations,
            "back_operations": back_operations
        }

    except Exception as e:
        logger.error(f"XLSXパースエラー: {e}")
        raise
